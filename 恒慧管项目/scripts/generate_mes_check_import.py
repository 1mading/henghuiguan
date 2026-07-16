# -*- coding: utf-8 -*-
"""生成 MES点检表1.0 统一导入 Excel（压缩版，≤15 条任务）"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "MES点检表1.0版本上线-统一导入.xlsx"

WYB = "王元斌 Martin"
HY = "黄艳-信息中心"
LY = "刘易"
WZW = "吴志伟"

PROJ = "MES点检表1.0版本上线"
DEPT = "实施交付部"
MGR = WYB
PDESC = "MES点检模块；初版已交付现场推进；目标2026-08-15验收关闭"
PSTART = "2026-06-01"
PEND = "2026-08-15"

HEADERS = [
    "项目名称", "所属部门", "项目负责人", "项目描述", "项目开始日期", "项目结束日期",
    "一级任务", "二级任务", "三级任务", "四级任务",
    "任务描述", "负责人", "协助人", "优先级", "预计工时", "计划开始日期", "截止日期", "任务类型",
]

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def row(l1, l2, desc, assignee, collab, priority, hours, start, end, task_type="常规"):
    return [
        PROJ, DEPT, MGR, PDESC, PSTART, PEND,
        l1, l2, "", "",
        desc, assignee, collab, priority, hours, start, end, task_type,
    ]


# 合并为 ≤15 条可执行任务（一级为阶段，二级为合并后的交付任务）
ROWS = [
    # —— 已完成留痕（3）——
    row(
        "已完成留痕", "需求整理与范围确认",
        "【已完成】业务需求、1.0边界、验收标准初稿",
        WYB, "", "重要", 24, "2026-06-01", "2026-06-18",
    ),
    row(
        "已完成留痕", "初版开发与交付包",
        "【已完成】核心功能开发、联调自测、交付包准备",
        WZW, WYB, "紧急", 52, "2026-06-10", "2026-07-10",
    ),
    row(
        "已完成留痕", "初版现场接收与启动推进",
        "【已完成】现场部署开通、对接产线班组启动试用",
        LY, HY, "紧急", 12, "2026-07-08", "2026-07-15",
    ),

    # —— 现场闭环（5）——
    row(
        "现场试运行与问题闭环", "试点推进与问题收集",
        "明确试点范围；日常督促执行；收集卡点/体验/数据问题；周同步",
        LY, HY, "紧急", 24, "2026-07-16", "2026-07-31",
    ),
    row(
        "现场试运行与问题闭环", "问题清单分级与冻结",
        "【M1·07-22】汇总台账、分级归类、冻结1.0必修项并确认排期责任",
        WYB, HY, "紧急", 8, "2026-07-18", "2026-07-23",
    ),
    row(
        "现场试运行与问题闭环", "缺陷修复与改版发布",
        "阻塞/重要问题修复，发布现场可用版本并出变更说明",
        WZW, WYB, "紧急", 23, "2026-07-20", "2026-07-31",
    ),
    row(
        "现场试运行与问题闭环", "现场回归与台账回写",
        "逐条验证必修项；关闭/重开/降级回写问题台账",
        LY, HY, "紧急", 11, "2026-07-23", "2026-08-02",
    ),
    row(
        "现场试运行与问题闭环", "试点稳定运行",
        "【M2·08-05】主流程连续跑通；完成率/异常抽查；确认可正式启用",
        LY, f"{WZW}、{HY}", "紧急", 14, "2026-07-25", "2026-08-05",
    ),

    # —— 培训上线（4）——
    row(
        "培训与正式启用", "操作指引与FAQ",
        "操作工/班长指引 + 常见问题FAQ",
        WYB, HY, "重要", 8, "2026-08-01", "2026-08-05",
    ),
    row(
        "培训与正式启用", "分角色培训",
        "场次安排、操作工/班长培训、答疑补训",
        LY, HY, "重要", 10, "2026-08-04", "2026-08-08",
    ),
    row(
        "培训与正式启用", "正式切换启用",
        "【M3·08-10】切换方案确认、生产环境检查、正式切换宣布与执行",
        LY, f"{WZW}、{HY}", "紧急", 7, "2026-08-07", "2026-08-10",
    ),
    row(
        "培训与正式启用", "上线首周支持",
        "日清分派、现场保障、紧急快修、首周运行小结",
        LY, f"{WZW}、{HY}", "紧急", 18, "2026-08-08", "2026-08-15",
    ),

    # —— 验收收尾（2）——
    row(
        "验收与收尾", "上线验收",
        "【M4·08-15】验收材料、对照标准走查、签字确认",
        WYB, f"{LY}、{HY}", "紧急", 6, "2026-08-10", "2026-08-15",
    ),
    row(
        "验收与收尾", "移交与总结",
        "1.1遗留清单、运维移交说明、项目总结纪要",
        WYB, HY, "普通", 6, "2026-08-13", "2026-08-15",
    ),

    # —— 临时（1）——
    row(
        "临时支持", "客户临时需求支持",
        "范围外诉求单独记工时；评估是否影响8/15验收",
        WYB, f"{WZW}、{LY}、{HY}", "紧急", 4, "2026-07-16", "2026-08-15", "临时",
    ),
]

assert len(ROWS) <= 15, len(ROWS)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "统一导入"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for r in ROWS:
        ws.append(r)
    for r in range(2, len(ROWS) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = BORDER

    widths = [22, 12, 16, 28, 12, 12, 20, 22, 8, 8, 40, 16, 22, 8, 8, 12, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note = wb.create_sheet("导入说明", 1)
    note["A1"] = "MES点检表1.0 · 压缩版任务（≤15条）"
    note["A1"].font = Font(bold=True, size=14, color="4F46E5")
    note.merge_cells("A1:B1")
    lines = [
        ("任务条数", str(len(ROWS))),
        ("合并原则", "同阶段、同负责人、同产出的细项合并为一条交付任务"),
        ("分工", "需求王元斌 / 开发吴志伟 / 现场刘易 / 跟进黄艳"),
        ("里程碑", "M1 07-22 清单冻结；M2 08-05 试点稳定；M3 08-10 正式启用；M4 08-15 验收"),
        ("导入", "恒慧管自动读第一张「统一导入」；已完成项导入后请手动标完成"),
        ("注意", "若已导入过旧版细任务，请先清理同名项目任务再导，或改用新项目名"),
    ]
    note["A2"] = "项"
    note["B2"] = "内容"
    note["A2"].fill = NOTE_FILL
    note["B2"].fill = NOTE_FILL
    for i, (k, v) in enumerate(lines, 3):
        note.cell(i, 1, k).font = Font(bold=True)
        note.cell(i, 2, v).alignment = WRAP
    note.column_dimensions["A"].width = 12
    note.column_dimensions["B"].width = 72

    wb.save(OUT)
    print(f"written: {OUT}")
    print(f"sheets: {wb.sheetnames}")
    print(f"rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
