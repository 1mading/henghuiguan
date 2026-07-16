# -*- coding: utf-8 -*-
"""Compare Excel V4 with embedded template fields in performance.js (text parse)."""
import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

ROOT = Path(r"e:\王元斌\钉钉小程序研发---王元斌\企业工单系统文件")
xlsx = ROOT / "2026实施交付部项目经理通用绩效考核表_V4.xlsx"
js_path = ROOT / "恒慧管项目" / "server" / "src" / "services" / "performance.js"

wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb.active
excel = []
section = None
for r in range(4, 22):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    if a and b is None and "部分" in str(a):
        section = a
        continue
    if a == "合计分数":
        break
    if b:
        excel.append(
            {
                "sectionTitle": section,
                "category": a,
                "title": b,
                "weight": ws.cell(r, 3).value,
                "maxScore": ws.cell(r, 4).value,
                "definition": ws.cell(r, 5).value,
                "target": ws.cell(r, 6).value,
                "scoringRule": ws.cell(r, 7).value,
                "dataSource": ws.cell(r, 9).value,
            }
        )

text = js_path.read_text(encoding="utf-8")
# extract indicators by finding objects with key:
blocks = re.findall(
    r"\{\s*key:\s*'([^']+)'(.*?)(?=\n\s*\{|\n\s*\],)",
    text,
    flags=re.S,
)

def field(block, name):
    m = re.search(rf"{name}:\s*(null|'((?:\\'|[^'])*)'|\"((?:\\\"|[^\"]*)*)\"|(-?\d+(?:\.\d+)?))", block)
    if not m:
        return None
    if m.group(1) == "null":
        return None
    if m.group(2) is not None:
        return m.group(2).replace("\\'", "'")
    if m.group(3) is not None:
        return m.group(3)
    return float(m.group(4)) if "." in m.group(4) else int(m.group(4))

inds = []
for key, block in blocks:
    if "section:" not in block:
        continue
    inds.append(
        {
            "key": key,
            "sectionTitle": field(block, "sectionTitle"),
            "category": field(block, "category"),
            "title": field(block, "title"),
            "weight": field(block, "weight"),
            "weightLabel": field(block, "weightLabel"),
            "maxScore": field(block, "maxScore"),
            "maxScoreLabel": field(block, "maxScoreLabel"),
            "definition": field(block, "definition"),
            "target": field(block, "target"),
            "scoringRule": field(block, "scoringRule"),
            "dataSource": field(block, "dataSource"),
        }
    )

name_m = re.search(r"name:\s*'([^']+)'", text)
remark_m = re.search(r"remark:\s*\n?\s*'([^']+)'", text)
print("TITLE excel:", ws.cell(1, 1).value)
print("TITLE tpl  :", name_m.group(1) if name_m else None)
print("remark excel:", ws.cell(23, 1).value)
print("remark tpl  :", remark_m.group(1) if remark_m else None)
print("excel n", len(excel), "tpl n", len(inds))

for i, (e, t) in enumerate(zip(excel, inds)):
    diffs = []
    for k in ["sectionTitle", "category", "title", "definition", "target", "scoringRule", "dataSource"]:
        ev, tv = e.get(k) or "", t.get(k) or ""
        if ev != tv:
            diffs.append(f"{k}:\n  E={ev!r}\n  T={tv!r}")
    ew = e["weight"]
    if ew in ("—", "—", "-"):
        if t.get("weight") is not None and t.get("weightLabel") not in ("—", "—"):
            diffs.append(f"weight E={ew!r} T={t.get('weight')!r} label={t.get('weightLabel')!r}")
    else:
        try:
            if float(ew) != float(t.get("weight")):
                diffs.append(f"weight E={ew!r} T={t.get('weight')!r}")
        except Exception:
            diffs.append(f"weight E={ew!r} T={t.get('weight')!r}")
    em = e["maxScore"]
    label = t.get("maxScoreLabel")
    val = t.get("maxScore")
    if str(em) != str(label) and str(em) != str(val):
        try:
            if float(em) != float(val):
                diffs.append(f"maxScore E={em!r} T={label!r}/{val!r}")
        except Exception:
            diffs.append(f"maxScore E={em!r} T={label!r}/{val!r}")
    if diffs:
        print(f"\nDIFF {i+1} {t.get('title')}:")
        for d in diffs:
            print(d)
    else:
        print(f"OK {i+1} {t.get('title')}")

print("\n=== Headers ===")
for c in range(1, 13):
    print(c, repr(ws.cell(3, c).value))

# column widths / row heights for layout fidelity
print("\n=== col widths ===")
for col in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
    dim = ws.column_dimensions[col]
    print(col, dim.width)

print("\nmerged", [str(m) for m in ws.merged_cells.ranges])
